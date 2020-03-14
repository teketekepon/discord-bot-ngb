# -*- coding: utf-8 -*-
import pyocr
import pyocr.builders
import pickle
import numpy as np
import aiohttp
import os
import re
import discord
from PIL import Image
from discord.ext import commands
from openpyxl import load_workbook

path_tesseract = r'./vendor/tesseract-ocr'
temp_path = r'./tmp/'
image_path = r'./downloads/image.png'
excel_path = r'./clan_battle_template.xlsx'
BOSSES = ['ワイバーン', 'ライライ', 'シードレイク', 'ムーバ', 'トルペドン']  # 2月のボス
STUMPS = ['△', '◆', '□', '◎', '☆', '〇']  # 左から[1ボスLA,2ボスLA,3ボスLA,4ボスLA,5ボスLA,凸]
work_channel_id = 641248473546489876  # バトルログのスクショを貼るチャンネルのID

class SaveResult(commands.Cog):
        # クラスのコンストラクタ。Botを受取り、インスタンス変数として保持。
    def __init__(self, bot):
        self.bot = bot
        if not os.path.isfile(temp_path + 'col.pkl'):
            self.col = [3] * 30
            self.rej = 9
            pickle.dump(self.col, open(temp_path + 'col.pkl','wb'))
            pickle.dump(self.rej, open(temp_path + 'rej.pkl','wb'))
        else:
            self.col = pickle.load(open(temp_path + 'col.pkl','rb'))
            self.rej = pickle.load(open(temp_path + 'rej.pkl','rb'))

    def cog_unload(self):
        pickle.dump(self.col, open(temp_path + 'col.pkl','wb'))
        pickle.dump(self.rej, open(temp_path + 'rej.pkl','wb'))

    async def download_img(self, url, file_name):
        chunk_size = 32
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as r:
                if r.status == 200:
                    with open(file_name, mode = 'wb') as fi:
                        while True:
                            chunk = await r.content.read(chunk_size)
                            if not chunk:
                                break
                            fi.write(chunk)
    # スクショからバトルログを抽出し2値化する関数
    def image_binarize(self, image):
        resolutions = [(1284, 715),
                                (1280, 720),
                                (1334, 750),
                                (1920, 1080),  # ここまで16:9
                                (2224, 1668),  # 4:3
                                (2880, 1440),  # 2_1a
                                (3040, 1440),  # 2_1a(左160黒い)
                                (2436, 1125),  # 2_1b
                                (2688, 1242)]  # 2_1b
        # resolutionsにある解像度なら読み取れる
        im = Image.open(image)
        for i, res in enumerate(resolutions):
            if im.size == res:
                num = i
                break
        if num is None:
            print('非対応の解像度です')
            return False
        if num <= 3:  # 16:9
            im_hd = im.resize((1920, 1080), Image.LANCZOS)
            im_crop = im_hd.crop((1400, 205, 1720, 920))
        elif num == 4:  # 4:3
            im_crop = im.crop((1625, 645, 1980, 1480))
        elif num == 5:  # 2_1a
            im_crop = im.crop((2200, 280, 2620, 1220))
        elif num == 6:  # 2_1a(左160黒い)
            im_crop = im.crop((2360, 280, 2780, 1220))
        else:  # 2_1b
            im_hd = im.resize((2668, 1242), Image.LANCZOS)
            im_crop = im_hd.crop((1965, 225, 2290, 1000))
        # numpyで2値化
        im_gray = np.array(im_crop.convert('L'))
        im_bin = (im_gray > 193) * 255
        # Save Binarized Image
        Image.fromarray(np.uint8(im_bin)).save(temp_path + 'temp.png')
        return True

    def useOCR(self, image):
        # ローカル環境の場合tesseractのpathを通す heroku環境の場合buildpackを使用するため不要
        # if path_tesseract not in os.environ["PATH"].split(os.pathsep):
        #     os.environ["PATH"] += os.pathsep + path_tesseract
        #     os.environ["TESSDATA_PREFIX"] = path_tesseract + r'/share/tessdata'
        tools = pyocr.get_available_tools()
        if len(tools) == 0:
            print('No OCR tool found')
            return
        tool = tools[0]
        # 日本語と英数字をOCR
        res = tool.image_to_string(Image.open(image),
                                   lang='jpn+eng',
                                   builder=pyocr.builders.WordBoxBuilder(tesseract_layout=6))
        text = ''
        for d in res:
            text = text + d.content
        # log/ocr_result.txtにocr結果を保存
        print(text)
        with open('./log/ocr_result.txt', mode = 'w', encoding = 'UTF-8') as f:
            f.write(text)
        return text

    def save_excel(self, text):
        workbook = load_workbook(excel_path)
        sheet = workbook['Battle_log']
        member_2l = [[cell.value for cell in tmp] for tmp in sheet['A2:A31']]
        member = sum(member_2l, [])  # 2次元配列なので1次元化
        data = re.findall(r'[グジで\\\w](.+?)が(.+?)に(.\d+)', text)  # 名前とボスとダメージのリスト
        if len(data) < 4:
            print('一部うまく読み取れなかったようです')
        # 凸かLAか判定するためのリスト('ダメージ'or'ダメージで'で判定するため'で'で始まる名前の人がいると使えません)
        isLA = re.findall(r'ダメージ.', text)
        for n, m in enumerate(data):  # nは添え字,mはタプル
            isMatch = False
            for i, j in enumerate(member):  # iは添え字,jはリスト
                if j is None:  # 空のセルは判定しない
                    break
                name_match = re.search(j, m[0])
                if name_match is None:
                    continue
                else:
                    isMatch = True
                row = i + 2  # cell[A2]から始まるため+2する
                if self.col[i] == self.rej:
                    print('1日6個以上のスタンプは押せません')
                    return
                if 'で' in isLA[n]:
                    for x, y in enumerate(BOSSES):  # LAなのでどのボスか判定
                        boss_match = re.search(y, m[1])
                        if boss_match is not None:
                            stu = x
                            break
                else:
                    stu = 5  # 凸なので〇スタンプをセット
                if  isMatch:
                    cell = sheet.cell(row=row, column=self.col[i], value=STUMPS[stu])
                    self.col[i] += 1
                continue
            if isMatch:
                print(f'{data[n]} は {cell} に\'{STUMPS[stu]}\'と書き込みました')
            else:
                print(f'{data[n]} はメンバーとマッチしなかった為書き込まれません')
        # Excelファイルをセーブして閉じる
        workbook.save(excel_path)
        workbook.close()

    # clearコマンドで利用する関数
    def clear_excel(self, kwd):
        # 任意のセルへ2次元配列を書き込む関数
        def write_list_2d(sheet, l_2d, start_row, start_col):
            for y, row in enumerate(l_2d):
                for x, cell in enumerate(row):
                    sheet.cell(row=start_row + y, column=start_col + x, value=l_2d[y][x])
        # 6*30の空の2次元配列を作る
        brank_list = [['' for row in range(6)] for col in range(30)]
        workbook = load_workbook(excel_path)
        sheet = workbook['Battle_log']
        if kwd == 'day1':  # 内容をクリア
            write_list_2d(sheet, brank_list, 2, 3)
        elif kwd == 'day2':
            write_list_2d(sheet, brank_list, 2, 10)
        elif kwd == 'day3':
            write_list_2d(sheet, brank_list, 2, 17)
        elif kwd == 'day4':
            write_list_2d(sheet, brank_list, 2, 24)
        elif kwd == 'day5':
            write_list_2d(sheet, brank_list, 2, 31)
        elif kwd == 'day6':
            write_list_2d(sheet, brank_list, 2, 38)
        elif kwd == 'all':
            for i in range(3, 39, 7):
                write_list_2d(sheet, brank_list, 2, i)
        else:
            print('clear_excel error: 無効な引数です')
        # Excelファイルをセーブして閉じる
        workbook.save(excel_path)
        workbook.close()

    @commands.group()  # 記録する位置を変更するコマンドグループ 全員の列位置が変更される
    async def day(self, ctx):
        if ctx.channel.id == work_channel_id:
            if ctx.invoked_subcommand is None:
                if self.rej == 9:
                    mes = '現在1日目です'
                elif self.rej == 16:
                    mes = '現在2日目です'
                elif self.rej == 23:
                    mes = '現在3日目です'
                elif self.rej == 30:
                    mes = '現在4日目です'
                elif self.rej == 37:
                    mes = '現在5日目です'
                else:
                    mes = '現在6日目です'
                await ctx.send(f'{mes} サブコマンドで何日目か教えてください ([1-6])')

    @day.command('1')
    async def day1(self, ctx):
        self.col = [3] * 30
        self.rej = 9
        await ctx.send('記録位置を1日目にセットしました')
    @day.command('2')
    async def day2(self, ctx):
        self.col = [10] * 30
        self.rej = 16
        await ctx.send('記録位置を2日目にセットしました')
    @day.command('3')
    async def day3(self, ctx):
        self.col = [17] * 30
        self.rej = 23
        await ctx.send('記録位置を3日目にセットしました')
    @day.command('4')
    async def day4(self, ctx):
        self.col = [24] * 30
        self.rej = 30
        await ctx.send('記録位置を4日目にセットしました')
    @day.command('5')
    async def day5(self, ctx):
        self.col = [31] * 30
        self.rej = 37
        await ctx.send('記録位置を5日目にセットしました')
    @day.command('6')
    async def day6(self, ctx):
        self.col = [38] * 30
        self.rej = 44
        await ctx.send('記録位置を6日目にセットしました')

    @commands.group()  # セルの内容を消去(Noneに上書き)するコマンドグループ
    async def clear(self, ctx):
        if ctx.channel.id == work_channel_id:
            if ctx.invoked_subcommand is None:
                await ctx.send('サブコマンドで何日目の記録を消去するか指定してください([1-6] または all)')

    @clear.command('1')
    async def clear_day1(self, ctx):
        self.clear_excel('day1')
        await ctx.send('1日目の記録内容を消去しました')
    @clear.command('2')
    async def clear_day2(self, ctx):
        self.clear_excel('day2')
        await ctx.send('2日目の記録内容を消去しました')
    @clear.command('3')
    async def clear_day3(self, ctx):
        self.clear_excel('day3')
        await ctx.send('3日目の記録内容を消去しました')
    @clear.command('4')
    async def clear_day4(self, ctx):
        self.clear_excel('day4')
        await ctx.send('4日目の記録内容を消去しました')
    @clear.command('5')
    async def clear_day5(self, ctx):
        self.clear_excel('day5')
        await ctx.send('5日目の記録内容を消去しました')
    @clear.command('6')
    async def clear_day6(self, ctx):
        self.clear_excel('day6')
        await ctx.send('6日目の記録内容を消去しました')
    @clear.command('all')
    async def clear_all(self, ctx):
        self.clear_excel('all')
        await ctx.send('すべての記録内容を消去しました')

    @commands.command()
    async def pull(self, ctx):
        await ctx.send(file=discord.File(excel_path))

    @commands.command('残凸')
    async def zantotu(self, ctx):
        workbook = load_workbook('./clan_battle_template.xlsx', data_only=True)
        sheet = workbook['Battle_log']
        for i in range(9, 45, 7):
            if i == self.rej:
                r = sheet.cell(row=32, column=i-2).value
        if r is not None:
            await ctx.send(f'残り凸数は {r} です')
        workbook.close()

    @commands.Cog.listener()
    async def on_message(self, message):
        if message.author.bot:
            return

        if message.channel.id == work_channel_id:
            if len(message.attachments) > 0:
                # messageに添付画像があり、指定のチャンネルの場合動作する
                await self.download_img(message.attachments[0].url, image_path)
                if self.image_binarize(image_path):
                    ocr_result = self.useOCR(temp_path + 'temp.png')
                if ocr_result is not None:
                    self.save_excel(ocr_result)
                # await message.channel.send(file=discord.File(temp_path + 'temp.png'))
# Bot本体側からコグを読み込む際に呼び出される関数。
def setup(bot):
    bot.add_cog(SaveResult(bot)) # クラスにBotを渡してインスタンス化し、Botにコグとして登録する。
