# -*- coding: utf-8 -*-
#
# バトルログのスクショを読み込みエクセルに結果を書き込む
#

import logging
from io import BytesIO
# import os
import re
import pickle
import typing

import discord
from discord.ext import commands
import jaconv
from openpyxl import load_workbook

from .lib.dbox import TransferData
from .lib.image_ocr import ImageOcr

TEMP_PATH = r'./tmp/'
EXCEL_PATH = r'./BattleLog.xlsx'
# 8月のボス
BOSSES = [
    'ワイバーン',
    'グリフォン',
    'マダムプリズム',
    'ムーバ',
    'メデューサ'
]
# 左から[1ボスLA,2ボスLA,3ボスLA,4ボスLA,5ボスLA,凸,不明]
STUMPS = ['△', '◆', '□', '◎', '☆', '〇', '?']


class SaveResult(commands.Cog):

    def __init__(self, bot):
        self.bot = bot
        self.logger = logging.getLogger('discord.SaveResult')
        # dict channels {key=channel.id value=[channel.name,rej,[col]]}
        # col: excel 列数値 rej: excel 書き込み禁止列数値(日にち判定に使用)
        self.channels = {}
        if TransferData().download_file(r'/channels.pkl',
                                        TEMP_PATH+'channels.pkl'):
            with open(TEMP_PATH + 'channels.pkl', 'rb') as f:
                self.channels = pickle.load(f)

    def cog_unload(self):
        with open(TEMP_PATH + 'channels.pkl', 'wb') as f:
            pickle.dump(self.channels, f)
        TransferData().upload_file(TEMP_PATH+'channels.pkl',
                                   r'/channels.pkl')

    def unvoiced(c):
        """清音にする"""
        def isHiragana(c):
            """cがひらがなのユニコードポイント範囲内か判定"""
            hiragana = range(0x3041, 0x3096 + 1)
            return ord(c) in hiragana

        def isKatakana(c):
            """cがカタカナのユニコードポイント範囲内か判定"""
            katakana = range(0x30A0, 0x30FA + 1)
            return ord(c) in katakana
        if isHiragana(c):              # ひらがなの場合
            hk = jaconv.hira2hkata(c)      # 半角ｶﾀｶﾅに変換
            zk = jaconv.h2z(hk[0])         # 一文字目だけを全角カタカナに戻す
            c = jaconv.kata2hira(zk)       # ひらがなに戻す
        elif isKatakana(c):            # カタカナの場合
            hk = jaconv.z2h(c)             # 半角ｶﾀｶﾅに変換
            c = jaconv.h2z(hk[0])          # 一文字目だけを全角カタカナに戻す
        return c

    def member_edit(self, sheet_name, op=None, *member):
        wb = load_workbook(EXCEL_PATH)
        sheet = wb[sheet_name]
        mlist = sum([[cell.value for cell in tmp] for tmp in sheet['A2:A31'
                                                                   ]], [])
        temp = []
        if op is None:
            wb.close()
            return ','.join(map(str, mlist))
        elif op == 'add':
            for m in member:
                if mlist not in member:
                    for i, item in enumerate(mlist):
                        if item is None:
                            mlist[i] = m
                            break
                temp.append(m)
        elif op == 'remove':
            for m in member:
                try:
                    mlist.remove(m)
                    mlist.append('')
                except ValueError:
                    temp.append(m)
        elif op == 'delete':
            if all(mlist):
                for i, item in enumerate(mlist):
                    if item is None:
                        mlist[i-1] = ''
                        break
            else:
                mlist[30] = ''
        elif op == 'clear':
            mlist = ['' for n in range(30)]
        else:
            wb.close()
            return ','.join(map(str, mlist))
        mlist = sorted(mlist, key=lambda x: (x is None, x))
        for num, w in enumerate(mlist):
            sheet.cell(row=num+2, column=1, value=w)
        wb.save(EXCEL_PATH)
        wb.close()
        return ','.join(map(str, mlist))

    def totu_count(self, text):
        n = 0
        # OCRtextから凸部分のみ抽出
        data = re.findall(r'ダメージで|ダメージ', text)
        if len(data) >= 5:
            del data[0]  # 1枚4件までのため
        for i in data:
            if 'で' not in i:
                n += 1
        return n

    def save_excel(self, sheet_name, rej, col,  text):
        workbook = load_workbook(EXCEL_PATH)
        sheet = workbook[sheet_name]
        # excelからメンバーリストを取得
        member = sum([[cell.value for cell in tmp] for tmp in sheet['A2:A31'
                                                                    ]], [])
        # 名前とボスとダメージのリスト抽出
        data = re.findall(r'(.+?)が(.+?)に(.[\doO]+)(ダメージで|ダメージ)', text)
        for n, m in enumerate(reversed(data)):  # nは添え字,mはタプル
            isMatch = False
            for i, j in enumerate(member):  # iは添え字,jはリスト
                if j is None:  # 空のセルは判定しない
                    break
                name_match = re.search(re.escape(j), m[0])
                if name_match is None:
                    continue
                else:
                    isMatch = True
                row = i + 2  # cell[A2]から始まるため+2する
                if col[i] == rej:
                    self.logger.info('%s 1日6個以上のスタンプは押せません', m[0])
                    return
                if 'で' in m[3]:
                    for x, y in enumerate(BOSSES):  # LAなのでどのボスか判定
                        boss_match = re.search(y, m[1])
                        if boss_match is not None:
                            stu = x
                            break
                    else:
                        stu = 5
                else:
                    stu = 5  # 凸なので〇スタンプをセット
                if isMatch:
                    sheet.cell(row=row, column=col[i], value=STUMPS[stu])
                    col[i] += 1
                    if stu == 5:
                        # self.channels[channel.id] += 1
                        pass
                continue
            if not isMatch:
                self.logger.info('%s はメンバーとマッチしなかった為書き込まれません', m[0])
        # Excelファイルをセーブして閉じる
        workbook.save(EXCEL_PATH)
        workbook.close()

    # clearコマンドで利用する関数
    def clear_excel(self, sheet_name, kwd):
        # 任意のセルを始点に2次元配列を書き込む関数
        def write_list_2d(sheet, l_2d, start_row, start_col):
            for y, row in enumerate(l_2d):
                for x, cell in enumerate(row):
                    sheet.cell(row=start_row + y,
                               column=start_col + x,
                               value=l_2d[y][x])
        # 6*30の空の2次元配列を作る
        blank_list = [['' for _ in range(6)] for _ in range(30)]
        wb = load_workbook(EXCEL_PATH)
        sheet = wb[sheet_name]
        if kwd == 'day1':  # 内容をクリア
            write_list_2d(sheet, blank_list, 2, 3)
        elif kwd == 'day2':
            write_list_2d(sheet, blank_list, 2, 10)
        elif kwd == 'day3':
            write_list_2d(sheet, blank_list, 2, 17)
        elif kwd == 'day4':
            write_list_2d(sheet, blank_list, 2, 24)
        elif kwd == 'day5':
            write_list_2d(sheet, blank_list, 2, 31)
        elif kwd == 'day6':
            write_list_2d(sheet, blank_list, 2, 38)
        elif kwd == 'all':
            for i in range(3, 39, 7):
                write_list_2d(sheet, blank_list, 2, i)
        else:
            self.logger.error('clear_excel: 無効な引数です')
            return False
        # Excelファイルをセーブして閉じる
        wb.save(EXCEL_PATH)
        wb.close()
        return True

    @commands.command()  # 記録する位置を変更するコマンド 全員の列位置が変更される
    async def day(self, ctx, a: typing.Optional[int] = None):
        if ctx.channel.id in self.channels.keys():
            var = self.channel[ctx.channels.id]
            if a is not None:
                if 0 < a < 6:
                    var[2] = [3+7*(a-1)] * 30
                    var[1] = 9+7*(a-1)
                    await ctx.send(f'記録位置を{a}日目にセットしました')
                else:
                    await ctx.send('引数は半角数字の1～6で入力してください')
            else:
                if var[1] == 9:
                    mes = '現在1日目です'
                elif var[1] == 16:
                    mes = '現在2日目です'
                elif var[1] == 23:
                    mes = '現在3日目です'
                elif var[1] == 30:
                    mes = '現在4日目です'
                elif var[1] == 37:
                    mes = '現在5日目です'
                elif var[1] == 44:
                    mes = '現在6日目です'
                else:
                    mes = '現在?日目です'
                await ctx.send(f'{mes} 引数で何日目か教えてください 1～6')
        else:
            await ctx.send('このチャンネルでの操作は許可されていません')

    @commands.command(name='del')  # セルの内容を消去(Noneに上書き)するコマンド
    async def _del(self, ctx, a: typing.Optional[str] = None):
        if ctx.channel.id in self.channels.keys():
            var = self.channel[ctx.channels.id]
            if a is not None:
                if len(a) < 2:
                    b = self.clear_excel(var[0], 'day' + a)
                else:
                    b = self.clear_excel(var[0], a)
                if b:
                    var[2] = [var[1] - 6] * 30
                    await ctx.send(f'「{a}」日目の内容を消去しました\nそれに伴い記録位置もリセットされました')
                else:
                    await ctx.send('引数が無効です')
            else:
                await ctx.send('何日目の記録を消去するか指定してください(1～6 または all)')
        else:
            await ctx.send('このチャンネルでの操作は許可されていません')

    @commands.command()
    async def member(self, ctx, op=None, *member):
        if ctx.channel.id in self.channels.keys():
            var = self.channel[ctx.channels.id]
            res = self.member_edit(var[0], op, *member)
            await ctx.send(f'現在のメンバー一覧です\n{res}')
        else:
            await ctx.send('このチャンネルでの操作は許可されていません')

    @commands.command()
    async def pull(self, ctx):
        if ctx.channel.id in self.channels.keys():
            await ctx.send(file=discord.File(EXCEL_PATH))

    # @commands.command()
    # async def totu(self, ctx):
    #     # await ctx.send(f'残り凸数は {90-self.channels} です')
    #     pass

    @commands.command()
    @commands.has_permissions(manage_channels=True)
    async def register(self, ctx):
        if ctx.channel.id in self.channels.keys():
            await ctx.send(f'{ctx.channel.name} はすでに作業チャンネルです')
        else:
            self.channels[ctx.channel.id] = [ctx.channel.name, 9, [3] * 30]
            wb = load_workbook(EXCEL_PATH)
            sheet = wb.copy_worksheet(wb['Template'])
            sheet.title = ctx.channel.name
            wb.save(EXCEL_PATH)
            wb.close()
            await ctx.send(f'{ctx.channel.name} を作業チャンネルに追加しました')

    @commands.command()
    @commands.has_permissions(manage_channels=True)
    async def unregister(self, ctx):
        if ctx.channel.id in self.channels.keys():
            wb = load_workbook(EXCEL_PATH)
            wb.remove_sheet(wb[self.channels[ctx.channel.id]])
            wb.save(EXCEL_PATH)
            wb.close()
            del self.channels[ctx.channel.id]
            await ctx.send(f'{ctx.channel.name} を作業チャンネルから除外しました')
        else:
            await ctx.send(f'{ctx.channel.name} は作業チャンネルではありません')

    @commands.Cog.listener()
    async def on_message(self, message):
        if message.author.bot:
            return

        if not message.attachments:
            return

        if message.channel.id in self.channels.keys():
            # messageに添付画像があり、指定のチャンネルの場合動作する
            image = BytesIO(await message.attachments[0].read())
            if (res := ImageOcr().image_ocr(image)) is not None:
                var = self.channel[message.channels.id]
                self.save_excel(var[0], var[1], var[2], res)


# Bot本体側からコグを読み込む際に呼び出される関数。
def setup(bot):
    bot.add_cog(SaveResult(bot))
